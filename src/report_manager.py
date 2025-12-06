# -*- coding: utf-8 -*-
"""
مدير التقارير لنظام إدارة المخزن
يقوم بإنشاء تقارير Excel و PDF
"""

import pandas as pd
import os
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import inch
import arabic_reshaper
from bidi.algorithm import get_display


class ReportManager:
    """مدير التقارير"""
    
    def __init__(self, excel_manager):
        self.excel_manager = excel_manager
        self.reports_path = "reports"
        os.makedirs(self.reports_path, exist_ok=True)
        
        # تسجيل الخط العربي
        try:
            # يمكنك إضافة خط عربي هنا إذا كان متاحاً
            # pdfmetrics.registerFont(TTFont('Arabic', 'path_to_arabic_font.ttf'))
            pass
        except:
            pass
    
    def reshape_arabic_text(self, text):
        """إعادة تشكيل النص العربي للعرض الصحيح في PDF"""
        try:
            if text and isinstance(text, str):
                reshaped_text = arabic_reshaper.reshape(text)
                bidi_text = get_display(reshaped_text)
                return bidi_text
            return str(text) if text is not None else ""
        except:
            return str(text) if text is not None else ""
    
    def export_inventory_to_excel(self, project_name):
        """تصدير المخزون الحالي إلى ملف Excel"""
        try:
            # الحصول على ملخص المخزون
            inventory_df = self.excel_manager.get_inventory_summary(project_name)
            
            if inventory_df.empty:
                return None, "لا توجد بيانات للتصدير"
            
            # إنشاء اسم الملف
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"مخزون_{project_name}_{timestamp}.xlsx"
            filepath = os.path.join(self.reports_path, filename)
            
            # تصدير إلى Excel مع تنسيق جميل
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # كتابة ملخص المخزون
                inventory_df.to_excel(writer, sheet_name='ملخص المخزون', index=False)
                
                # الحصول على الحركات التفصيلية
                project_file = os.path.join("projects", f"{project_name}_Transactions.xlsx")
                if os.path.exists(project_file):
                    transactions_df = pd.read_excel(project_file, engine='openpyxl')
                    transactions_df.to_excel(writer, sheet_name='الحركات التفصيلية', index=False)
                
                # تنسيق الأوراق
                workbook = writer.book
                
                # تنسيق ورقة المخزون
                inventory_sheet = writer.sheets['ملخص المخزون']
                self._format_excel_sheet(inventory_sheet, workbook)
                
                # تنسيق ورقة الحركات إذا وجدت
                if 'الحركات التفصيلية' in writer.sheets:
                    transactions_sheet = writer.sheets['الحركات التفصيلية']
                    self._format_excel_sheet(transactions_sheet, workbook)
            
            return filepath, "تم تصدير التقرير بنجاح"
            
        except Exception as e:
            return None, f"خطأ في تصدير التقرير: {str(e)}"
    
    def _format_excel_sheet(self, sheet, workbook):
        """تنسيق ورقة Excel"""
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        # تنسيق الرؤوس
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # تطبيق التنسيق على الصف الأول (الرؤوس)
        for cell in sheet[1]:
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # تنسيق البيانات
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal="center", vertical="center")
        
        # تطبيق التنسيق على باقي الصفوف
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_alignment
        
        # ضبط عرض الأعمدة
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def export_inventory_to_pdf(self, project_name):
        """تصدير المخزون الحالي إلى ملف PDF"""
        try:
            # الحصول على ملخص المخزون
            inventory_df = self.excel_manager.get_inventory_summary(project_name)
            
            if inventory_df.empty:
                return None, "لا توجد بيانات للتصدير"
            
            # إنشاء اسم الملف
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"مخزون_{project_name}_{timestamp}.pdf"
            filepath = os.path.join(self.reports_path, filename)
            
            # إنشاء مستند PDF
            doc = SimpleDocTemplate(
                filepath, 
                pagesize=landscape(A4),
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=1*inch,
                bottomMargin=0.5*inch
            )
            
            # محتويات المستند
            story = []
            
            # إضافة العنوان
            title_style = ParagraphStyle(
                'Title',
                fontSize=16,
                spaceAfter=20,
                alignment=1,  # توسيط
                fontName='Helvetica-Bold'
            )
            
            title_text = self.reshape_arabic_text(f"تقرير مخزون مشروع: {project_name}")
            title = Paragraph(title_text, title_style)
            story.append(title)
            
            # إضافة التاريخ
            date_text = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            date_text = self.reshape_arabic_text(date_text)
            date_style = ParagraphStyle('Date', fontSize=12, spaceAfter=20, alignment=1)
            date_para = Paragraph(date_text, date_style)
            story.append(date_para)
            
            story.append(Spacer(1, 20))
            
            # إنشاء الجدول
            table_data = []
            
            # إضافة الرؤوس
            headers = [
                self.reshape_arabic_text('اسم العنصر'),
                self.reshape_arabic_text('التصنيف'),
                self.reshape_arabic_text('الكمية الحالية'),
                self.reshape_arabic_text('إجمالي الداخل'),
                self.reshape_arabic_text('إجمالي الخارج'),
                self.reshape_arabic_text('مدة الصلاحية')
            ]
            table_data.append(headers)
            
            # إضافة البيانات
            for _, row in inventory_df.iterrows():
                row_data = [
                    self.reshape_arabic_text(str(row['اسم_العنصر'])),
                    self.reshape_arabic_text(str(row['التصنيف'])),
                    str(row['الكمية_الحالية']),
                    str(row['إجمالي_الداخل']),
                    str(row['إجمالي_الخارج']),
                    str(row['مدة_الصلاحية_بالأيام']) if pd.notna(row['مدة_الصلاحية_بالأيام']) else '-'
                ]
                table_data.append(row_data)
            
            # إنشاء الجدول
            table = Table(table_data, repeatRows=1)
            
            # تنسيق الجدول
            table.setStyle(TableStyle([
                # تنسيق الرؤوس
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                
                # تنسيق البيانات
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # ألوان متناوبة للصفوف
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            story.append(table)
            
            # بناء المستند
            doc.build(story)
            
            return filepath, "تم إنشاء تقرير PDF بنجاح"
            
        except Exception as e:
            return None, f"خطأ في إنشاء تقرير PDF: {str(e)}"
    
    def print_report(self, filepath):
        """طباعة التقرير"""
        try:
            import subprocess
            import sys
            
            if sys.platform == "win32":
                # على نظام Windows
                os.startfile(filepath, "print")
                return True, "تم إرسال الملف للطباعة"
            else:
                # على أنظمة أخرى
                subprocess.run(["lpr", filepath])
                return True, "تم إرسال الملف للطباعة"
                
        except Exception as e:
            return False, f"خطأ في الطباعة: {str(e)}"
    
    def get_project_statistics(self, project_name):
        """الحصول على إحصائيات المشروع"""
        try:
            project_file = os.path.join("projects", f"{project_name}_Transactions.xlsx")
            
            if not os.path.exists(project_file):
                return {}
            
            df = pd.read_excel(project_file, engine='openpyxl')
            
            if df.empty:
                return {}
            
            stats = {
                'إجمالي_الحركات': len(df),
                'عدد_حركات_الدخول': len(df[df['نوع_العملية'] == 'دخول']),
                'عدد_حركات_الخروج': len(df[df['نوع_العملية'] == 'خروج']),
                'عدد_العناصر_المختلفة': df['اسم_العنصر'].nunique(),
                'التصنيفات_المختلفة': df['التصنيف'].nunique(),
                'آخر_حركة': df['التاريخ'].max() if not df.empty else None
            }
            
            return stats
            
        except Exception as e:
            print(f"خطأ في حساب الإحصائيات: {e}")
            return {}
    
    def export_comprehensive_report_to_excel(self, project_name):
        """تصدير تقرير شامل إلى Excel يتضمن جميع البيانات والتواريخ"""
        try:
            # إنشاء اسم الملف
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"تقرير_شامل_{project_name}_{timestamp}.xlsx"
            filepath = os.path.join(self.reports_path, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                
                # 1. ملخص المخزون الحالي
                inventory_df = self.excel_manager.get_inventory_summary(project_name)
                if not inventory_df.empty:
                    inventory_df.to_excel(writer, sheet_name='المخزون الحالي', index=False)
                
                # 2. جميع الحركات التفصيلية
                project_file = os.path.join("projects", f"{project_name}_Transactions.xlsx")
                if os.path.exists(project_file):
                    transactions_df = pd.read_excel(project_file, engine='openpyxl')
                    
                    # ترتيب الحركات حسب التاريخ (الأحدث أولاً)
                    transactions_df['التاريخ'] = pd.to_datetime(transactions_df['التاريخ'])
                    transactions_df = transactions_df.sort_values('التاريخ', ascending=False)
                    
                    # تنسيق التاريخ للعرض
                    transactions_df['التاريخ'] = transactions_df['التاريخ'].dt.strftime('%Y-%m-%d %H:%M:%S')
                    
                    transactions_df.to_excel(writer, sheet_name='جميع الحركات', index=False)
                
                # 3. تقرير حركات الدخول
                if os.path.exists(project_file):
                    incoming_df = transactions_df[transactions_df['نوع_العملية'] == 'دخول'].copy()
                    if not incoming_df.empty:
                        incoming_df.to_excel(writer, sheet_name='حركات الدخول', index=False)
                
                # 4. تقرير حركات الخروج
                if os.path.exists(project_file):
                    outgoing_df = transactions_df[transactions_df['نوع_العملية'] == 'خروج'].copy()
                    if not outgoing_df.empty:
                        outgoing_df.to_excel(writer, sheet_name='حركات الخروج', index=False)
                
                # 5. إحصائيات المشروع
                stats = self.get_project_statistics(project_name)
                if stats:
                    stats_df = pd.DataFrame([stats]).T
                    stats_df.columns = ['القيمة']
                    stats_df.index.name = 'الإحصائية'
                    stats_df.to_excel(writer, sheet_name='إحصائيات المشروع')
                
                # 6. ملخص العناصر الأساسية
                master_items = self.excel_manager.get_all_items()
                if not master_items.empty:
                    master_items.to_excel(writer, sheet_name='العناصر الأساسية', index=False)
                
                # 7. تقرير منفصل حسب التصنيف
                if not inventory_df.empty:
                    categories = inventory_df['التصنيف'].unique()
                    for category in categories:
                        category_data = inventory_df[inventory_df['التصنيف'] == category]
                        sheet_name = f"تصنيف_{category}"[:31]  # Excel sheet name limit
                        category_data.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # تنسيق جميع الأوراق
                workbook = writer.book
                for sheet_name in writer.sheets:
                    sheet = writer.sheets[sheet_name]
                    self._format_excel_sheet(sheet, workbook)
            
            return filepath, "تم إنشاء التقرير الشامل بنجاح"
            
        except Exception as e:
            return None, f"خطأ في إنشاء التقرير الشامل: {str(e)}"
    
    def export_filtered_report_to_excel(self, project_name, start_date, end_date):
        """تصدير تقرير مفلتر بالتاريخ إلى Excel"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"تقرير_مفلتر_{project_name}_{timestamp}.xlsx"
            filepath = os.path.join(self.reports_path, filename)
            
            # تحويل التواريخ
            if isinstance(start_date, str):
                start_date = pd.to_datetime(start_date)
            elif hasattr(start_date, 'date') and callable(start_date.date):
                start_date = pd.to_datetime(start_date.date())
            elif not isinstance(start_date, (pd.Timestamp, datetime)):
                start_date = pd.to_datetime(start_date)
            
            if isinstance(end_date, str):
                end_date = pd.to_datetime(end_date)
            elif hasattr(end_date, 'date') and callable(end_date.date):
                end_date = pd.to_datetime(end_date.date())
            elif not isinstance(end_date, (pd.Timestamp, datetime)):
                end_date = pd.to_datetime(end_date)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                
                # 1. معلومات الفلتر
                filter_info = {
                    'معلومات التقرير': [
                        f"اسم المشروع: {project_name}",
                        f"من تاريخ: {start_date.strftime('%Y-%m-%d')}",
                        f"إلى تاريخ: {end_date.strftime('%Y-%m-%d')}",
                        f"تاريخ إنشاء التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                        f"عدد الأيام: {(end_date - start_date).days + 1}"
                    ]
                }
                filter_df = pd.DataFrame(filter_info)
                filter_df.to_excel(writer, sheet_name='معلومات التقرير', index=False)
                
                # قراءة الحركات المفلترة
                project_file = os.path.join("projects", f"{project_name}_Transactions.xlsx")
                filtered_transactions = pd.DataFrame()
                
                if os.path.exists(project_file):
                    all_transactions = pd.read_excel(project_file, engine='openpyxl')
                    
                    if not all_transactions.empty:
                        # تحويل عمود التاريخ
                        all_transactions['التاريخ'] = pd.to_datetime(all_transactions['التاريخ'], errors='coerce')
                        
                        # إزالة الصفوف التي لا تحتوي على تواريخ صحيحة
                        all_transactions = all_transactions.dropna(subset=['التاريخ'])
                        
                        # تطبيق الفلتر
                        filtered_transactions = all_transactions[
                            (all_transactions['التاريخ'] >= start_date) &
                            (all_transactions['التاريخ'] <= end_date)
                        ].copy()
                        
                        if not filtered_transactions.empty:
                            # ترتيب الحركات حسب التاريخ (الأحدث أولاً)
                            filtered_transactions = filtered_transactions.sort_values('التاريخ', ascending=False)
                            
                            # تنسيق التاريخ للعرض
                            filtered_transactions['التاريخ'] = filtered_transactions['التاريخ'].dt.strftime('%Y-%m-%d %H:%M:%S')
                
                # 2. جميع الحركات المفلترة
                if not filtered_transactions.empty:
                    filtered_transactions.to_excel(writer, sheet_name='الحركات المفلترة', index=False)
                else:
                    empty_df = pd.DataFrame({'رسالة': ['لا توجد حركات في الفترة المحددة']})
                    empty_df.to_excel(writer, sheet_name='الحركات المفلترة', index=False)
                
                # 3. حركات الدخول المفلترة
                if not filtered_transactions.empty:
                    incoming_filtered = filtered_transactions[filtered_transactions['نوع_العملية'] == 'دخول'].copy()
                    if not incoming_filtered.empty:
                        incoming_filtered.to_excel(writer, sheet_name='دخول مفلتر', index=False)
                
                # 4. حركات الخروج المفلترة
                if not filtered_transactions.empty:
                    outgoing_filtered = filtered_transactions[filtered_transactions['نوع_العملية'] == 'خروج'].copy()
                    if not outgoing_filtered.empty:
                        outgoing_filtered.to_excel(writer, sheet_name='خروج مفلتر', index=False)
                
                # 5. إحصائيات الفترة المحددة
                if not filtered_transactions.empty:
                    period_stats = self._calculate_period_statistics(filtered_transactions, start_date, end_date)
                    if period_stats:
                        stats_df = pd.DataFrame([period_stats]).T
                        stats_df.columns = ['القيمة']
                        stats_df.index.name = 'الإحصائية'
                        stats_df.to_excel(writer, sheet_name='إحصائيات الفترة')
                
                # 6. ملخص العناصر المتأثرة في الفترة
                if not filtered_transactions.empty:
                    affected_items = self._get_affected_items_summary(filtered_transactions)
                    if not affected_items.empty:
                        affected_items.to_excel(writer, sheet_name='العناصر المتأثرة', index=False)
                
                # 7. المخزون الحالي (مرجعي)
                current_inventory = self.excel_manager.get_current_inventory(project_name)
                if not current_inventory.empty:
                    current_inventory.to_excel(writer, sheet_name='المخزون الحالي', index=False)
                
                # تنسيق جميع الأوراق
                workbook = writer.book
                for sheet_name in writer.sheets:
                    sheet = writer.sheets[sheet_name]
                    self._format_excel_sheet(sheet, workbook)
            
            return filepath, "تم إنشاء التقرير المفلتر بنجاح"
            
        except Exception as e:
            return None, f"خطأ في إنشاء التقرير المفلتر: {str(e)}"
    
    def _calculate_period_statistics(self, transactions_df, start_date, end_date):
        """حساب إحصائيات فترة زمنية محددة"""
        try:
            stats = {}
            
            # إحصائيات عامة
            stats['إجمالي الحركات'] = len(transactions_df)
            stats['عدد حركات الدخول'] = len(transactions_df[transactions_df['نوع_العملية'] == 'دخول'])
            stats['عدد حركات الخروج'] = len(transactions_df[transactions_df['نوع_العملية'] == 'خروج'])
            
            # إحصائيات الكميات
            incoming_qty = transactions_df[transactions_df['نوع_العملية'] == 'دخول']['الكمية'].sum()
            outgoing_qty = transactions_df[transactions_df['نوع_العملية'] == 'خروج']['الكمية'].sum()
            
            stats['إجمالي كميات الدخول'] = f"{incoming_qty:,.0f}"
            stats['إجمالي كميات الخروج'] = f"{outgoing_qty:,.0f}"
            stats['صافي الحركة'] = f"{incoming_qty - outgoing_qty:,.0f}"
            
            # العناصر المتأثرة
            unique_items = transactions_df['اسم_العنصر'].nunique()
            stats['عدد العناصر المتأثرة'] = unique_items
            
            # التصنيفات المتأثرة
            unique_categories = transactions_df['التصنيف'].nunique()
            stats['عدد التصنيفات المتأثرة'] = unique_categories
            
            # معدلات يومية
            period_days = (end_date - start_date).days + 1
            stats['متوسط الحركات اليومية'] = f"{len(transactions_df) / period_days:.1f}"
            stats['متوسط كمية الدخول اليومية'] = f"{incoming_qty / period_days:,.1f}"
            stats['متوسط كمية الخروج اليومية'] = f"{outgoing_qty / period_days:,.1f}"
            
            # معلومات الفترة
            stats['فترة التقرير من'] = start_date.strftime('%Y-%m-%d')
            stats['فترة التقرير إلى'] = end_date.strftime('%Y-%m-%d')
            stats['عدد أيام الفترة'] = period_days
            
            return stats
            
        except Exception as e:
            return {'خطأ': f'لا يمكن حساب الإحصائيات: {str(e)}'}
    
    def _get_affected_items_summary(self, transactions_df):
        """الحصول على ملخص العناصر المتأثرة في الفترة"""
        try:
            # تجميع البيانات حسب العنصر
            summary = transactions_df.groupby(['اسم_العنصر', 'التصنيف']).agg({
                'الكمية': ['count', 'sum'],
                'نوع_العملية': lambda x: ', '.join(x.unique()),
                'التاريخ': ['min', 'max']
            }).round(2)
            
            # تسطيح أسماء الأعمدة
            summary.columns = ['عدد الحركات', 'إجمالي الكمية', 'أنواع العمليات', 'أول حركة', 'آخر حركة']
            summary = summary.reset_index()
            
            # ترتيب حسب إجمالي الكمية
            summary = summary.sort_values('إجمالي الكمية', ascending=False)
            
            return summary
            
        except Exception as e:
            return pd.DataFrame({'خطأ': [f'لا يمكن إنشاء ملخص العناصر: {str(e)}']})
    
    def export_ultra_comprehensive_report(self, project_name):
        """تصدير تقرير شامل ومفصل يحتوي على كل التفاصيل الممكنة"""
        try:
            # إنشاء اسم الملف
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"تقرير_شامل_مفصل_{project_name}_{timestamp}.xlsx"
            filepath = os.path.join(self.reports_path, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                
                # 1. معلومات عامة عن التقرير
                report_info = {
                    'معلومات التقرير': [
                        f"اسم المشروع: {project_name}",
                        f"تاريخ إنشاء التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                        f"نوع التقرير: تقرير شامل ومفصل",
                        f"وصف: تقرير يحتوي على جميع بيانات المشروع وإحصائياته",
                        f"الإصدار: نظام إدارة المخزن - StrucTech"
                    ]
                }
                info_df = pd.DataFrame(report_info)
                info_df.to_excel(writer, sheet_name='معلومات التقرير', index=False)
                
                # 2. الملخص التنفيذي (للإدارة)
                executive_summary = self._create_executive_summary(project_name)
                if executive_summary is not None and not executive_summary.empty:
                    executive_summary.to_excel(writer, sheet_name='الملخص التنفيذي', index=False)
                
                # 3. المخزون الحالي (مع التفاصيل)
                try:
                    inventory_df = self.excel_manager.get_inventory_summary(project_name)
                    if inventory_df is not None and not inventory_df.empty:
                        # إضافة تفاصيل إضافية
                        enhanced_inventory = self._enhance_inventory_details(inventory_df)
                        enhanced_inventory.to_excel(writer, sheet_name='المخزون الحالي', index=False)
                except Exception as e:
                    print(f"خطأ في المخزون الحالي: {e}")
                
                # 4. جميع الحركات (مرتبة بالتاريخ)
                try:
                    project_file = os.path.join("projects", f"{project_name}_Transactions.xlsx")
                    if os.path.exists(project_file):
                        all_transactions = pd.read_excel(project_file, engine='openpyxl')
                        if all_transactions is not None and not all_transactions.empty:
                            # ترتيب وتحسين الحركات
                            enhanced_transactions = self._enhance_transactions_details(all_transactions)
                            enhanced_transactions.to_excel(writer, sheet_name='جميع الحركات', index=False)
                except Exception as e:
                    print(f"خطأ في جميع الحركات: {e}")
                    enhanced_transactions = pd.DataFrame()  # DataFrame فارغ للاستخدام لاحقاً
                
                # معالجة الحركات إذا كانت متوفرة
                if 'enhanced_transactions' in locals() and enhanced_transactions is not None and not enhanced_transactions.empty:
                    # 5. حركات الدخول (مفصلة)
                    try:
                        incoming_detailed = enhanced_transactions[enhanced_transactions['نوع_العملية'] == 'إدخال'].copy()
                        if not incoming_detailed.empty:
                            incoming_detailed.to_excel(writer, sheet_name='حركات الدخول', index=False)
                    except Exception as e:
                        print(f"خطأ في حركات الدخول: {e}")
                    
                    # 6. حركات الخروج (مفصلة)
                    try:
                        outgoing_detailed = enhanced_transactions[enhanced_transactions['نوع_العملية'] == 'إخراج'].copy()
                        if not outgoing_detailed.empty:
                            outgoing_detailed.to_excel(writer, sheet_name='حركات الخروج', index=False)
                    except Exception as e:
                        print(f"خطأ في حركات الخروج: {e}")
                    
                    # 7. تحليل حسب التصنيف
                    try:
                        category_analysis = self._analyze_by_category(enhanced_transactions)
                        if category_analysis is not None and not category_analysis.empty:
                            category_analysis.to_excel(writer, sheet_name='تحليل التصنيفات', index=False)
                    except Exception as e:
                        print(f"خطأ في تحليل التصنيفات: {e}")
                    
                    # 8. تحليل زمني (شهري)
                    try:
                        time_analysis = self._analyze_by_time(enhanced_transactions)
                        if time_analysis is not None and not time_analysis.empty:
                            time_analysis.to_excel(writer, sheet_name='التحليل الزمني', index=False)
                    except Exception as e:
                        print(f"خطأ في التحليل الزمني: {e}")
                    
                    # 9. أعلى العناصر حركة
                    try:
                        top_items = self._get_top_items_analysis(enhanced_transactions)
                        if top_items is not None and not top_items.empty:
                            top_items.to_excel(writer, sheet_name='أعلى العناصر حركة', index=False)
                    except Exception as e:
                        print(f"خطأ في تحليل أعلى العناصر: {e}")
                
                # 10. العناصر الأساسية (قاعدة البيانات)
                try:
                    master_items = self.excel_manager.get_all_items()
                    if master_items is not None and not master_items.empty:
                        master_items.to_excel(writer, sheet_name='قاعدة العناصر', index=False)
                except Exception as e:
                    print(f"خطأ في قاعدة العناصر: {e}")
                
                # 11. تنبيهات المخزون والصلاحية
                try:
                    alerts_data = self._get_alerts_data(project_name)
                    if alerts_data and len(alerts_data) > 0:
                        alerts_df = pd.DataFrame(alerts_data)
                        alerts_df.to_excel(writer, sheet_name='تنبيهات وإنذارات', index=False)
                except Exception as e:
                    print(f"خطأ في التنبيهات: {e}")
                
                # 12. إحصائيات شاملة
                try:
                    comprehensive_stats = self._get_comprehensive_statistics(project_name)
                    if comprehensive_stats and len(comprehensive_stats) > 0:
                        stats_df = pd.DataFrame([comprehensive_stats]).T
                        stats_df.columns = ['القيمة']
                        stats_df.index.name = 'الإحصائية'
                        stats_df.to_excel(writer, sheet_name='إحصائيات شاملة')
                except Exception as e:
                    print(f"خطأ في الإحصائيات: {e}")
                
                # تنسيق جميع الأوراق
                workbook = writer.book
                for sheet_name in writer.sheets:
                    sheet = writer.sheets[sheet_name]
                    self._format_excel_sheet(sheet, workbook)
            
            return filepath, "تم إنشاء التقرير الشامل بنجاح! يحتوي على 12 ورقة عمل مع جميع التفاصيل"
            
        except Exception as e:
            return None, f"خطأ في إنشاء التقرير الشامل: {str(e)}"
    
    def _create_executive_summary(self, project_name):
        """إنشاء الملخص التنفيذي"""
        try:
            # الحصول على البيانات الأساسية
            inventory_df = self.excel_manager.get_inventory_summary(project_name)
            
            # إحصائيات أساسية
            total_items = len(inventory_df) if not inventory_df.empty else 0
            total_quantity = inventory_df['الكمية_الحالية'].sum() if not inventory_df.empty else 0
            
            # حالة المخزون
            low_stock_items = 0
            if not inventory_df.empty:
                low_stock_items = len(inventory_df[inventory_df['الكمية_الحالية'] <= 10])
            
            # إحصائيات الحركات
            project_file = os.path.join("projects", f"{project_name}_Transactions.xlsx")
            transactions_count = 0
            latest_transaction = "غير متوفر"
            
            if os.path.exists(project_file):
                transactions_df = pd.read_excel(project_file, engine='openpyxl')
                transactions_count = len(transactions_df)
                if not transactions_df.empty:
                    # تحويل التاريخ بأمان
                    try:
                        transactions_df['التاريخ'] = pd.to_datetime(transactions_df['التاريخ'])
                        latest_transaction = transactions_df['التاريخ'].max().strftime('%Y/%m/%d')
                    except:
                        latest_transaction = str(transactions_df['التاريخ'].max())
            
            # إنشاء الملخص
            summary_data = {
                'البيان': [
                    'إجمالي العناصر المسجلة',
                    'إجمالي الكميات الحالية', 
                    'عناصر بمخزون منخفض',
                    'إجمالي الحركات المسجلة',
                    'آخر حركة مسجلة',
                    'تاريخ إنشاء التقرير',
                    'حالة المشروع'
                ],
                'القيمة': [
                    total_items,
                    int(total_quantity),
                    low_stock_items,
                    transactions_count,
                    latest_transaction,
                    datetime.now().strftime('%Y/%m/%d %H:%M'),
                    'نشط'
                ]
            }
            
            return pd.DataFrame(summary_data)
            
        except Exception as e:
            print(f"خطأ في إنشاء الملخص التنفيذي: {e}")
            return pd.DataFrame({'البيان': ['خطأ'], 'القيمة': ['فشل في إنشاء الملخص']})
    
    def _enhance_inventory_details(self, inventory_df):
        """تحسين تفاصيل المخزون"""
        try:
            if inventory_df is None or inventory_df.empty:
                return inventory_df
                
            enhanced_df = inventory_df.copy()
            
            # إضافة حالة المخزون
            enhanced_df['حالة_المخزون'] = enhanced_df['الكمية_الحالية'].apply(
                lambda x: 'مخزون منخفض' if x <= 10 else 'مخزون مناسب' if x <= 50 else 'مخزون مرتفع'
            )
            
            # إضافة نسبة الاستخدام التقريبية
            enhanced_df['تقييم_الاستخدام'] = enhanced_df['الكمية_الحالية'].apply(
                lambda x: 'عالي الاستخدام' if x <= 20 else 'متوسط الاستخدام' if x <= 100 else 'قليل الاستخدام'
            )
            
            return enhanced_df
            
        except Exception as e:
            print(f"خطأ في تحسين بيانات المخزون: {e}")
            return inventory_df
    
    def _analyze_by_category(self, transactions_df):
        """تحليل الحركات حسب التصنيف"""
        try:
            if transactions_df is None or transactions_df.empty:
                return pd.DataFrame()
            
            # التحقق من وجود الأعمدة المطلوبة
            required_columns = ['التصنيف', 'نوع_العملية', 'الكمية', 'التاريخ']
            missing_columns = [col for col in required_columns if col not in transactions_df.columns]
            if missing_columns:
                print(f"أعمدة مفقودة في تحليل التصنيفات: {missing_columns}")
                return pd.DataFrame()
            
            # تجميع حسب التصنيف ونوع العملية
            category_analysis = transactions_df.groupby(['التصنيف', 'نوع_العملية']).agg({
                'الكمية': 'sum',
                'التاريخ': 'count'
            }).reset_index()
            
            category_analysis.columns = ['التصنيف', 'نوع_العملية', 'إجمالي_الكمية', 'عدد_العمليات']
            
            return category_analysis
            
        except Exception as e:
            print(f"خطأ في تحليل التصنيفات: {e}")
            return pd.DataFrame()
    
    def _analyze_by_time(self, transactions_df):
        """تحليل الحركات حسب الوقت"""
        try:
            if transactions_df is None or transactions_df.empty:
                return pd.DataFrame()
            
            # التحقق من وجود الأعمدة المطلوبة
            required_columns = ['التاريخ', 'نوع_العملية', 'الكمية']
            missing_columns = [col for col in required_columns if col not in transactions_df.columns]
            if missing_columns:
                print(f"أعمدة مفقودة في التحليل الزمني: {missing_columns}")
                return pd.DataFrame()
            
            # عمل نسخة لتجنب تعديل الأصل
            df_copy = transactions_df.copy()
            
            # تحويل التاريخ
            df_copy['التاريخ'] = pd.to_datetime(df_copy['التاريخ'])
            df_copy['الشهر'] = df_copy['التاريخ'].dt.strftime('%Y-%m')
            
            # تحليل شهري
            monthly_analysis = df_copy.groupby(['الشهر', 'نوع_العملية']).agg({
                'الكمية': 'sum',
                'التاريخ': 'count'
            }).reset_index()
            
            monthly_analysis.columns = ['الشهر', 'نوع_العملية', 'إجمالي_الكمية', 'عدد_العمليات']
            
            return monthly_analysis
            
        except Exception as e:
            print(f"خطأ في التحليل الزمني: {e}")
            return pd.DataFrame()
    
    def _get_top_items_analysis(self, transactions_df):
        """تحليل أعلى العناصر حركة"""
        try:
            if transactions_df is None or transactions_df.empty:
                return pd.DataFrame()
            
            # التحقق من وجود الأعمدة المطلوبة
            required_columns = ['اسم_العنصر', 'الكمية', 'التاريخ', 'نوع_العملية']
            missing_columns = [col for col in required_columns if col not in transactions_df.columns]
            if missing_columns:
                print(f"أعمدة مفقودة في تحليل أعلى العناصر: {missing_columns}")
                return pd.DataFrame()
            
            # تجميع حسب العنصر
            item_analysis = transactions_df.groupby('اسم_العنصر').agg({
                'الكمية': 'sum',
                'التاريخ': 'count',
                'نوع_العملية': lambda x: ', '.join(x.unique())
            }).reset_index()
            
            item_analysis.columns = ['اسم_العنصر', 'إجمالي_الكمية', 'عدد_العمليات', 'أنواع_العمليات']
            
            # ترتيب حسب العدد
            item_analysis = item_analysis.sort_values('عدد_العمليات', ascending=False).head(20)
            
            return item_analysis
            
        except Exception as e:
            print(f"خطأ في تحليل أعلى العناصر: {e}")
            return pd.DataFrame()
    
    def _get_alerts_data(self, project_name):
        """الحصول على بيانات التنبيهات"""
        try:
            alerts = []
            
            # تنبيهات المخزون المنخفض
            inventory_df = self.excel_manager.get_inventory_summary(project_name)
            if not inventory_df.empty:
                low_stock = inventory_df[inventory_df['الكمية_الحالية'] <= 10]
                for _, item in low_stock.iterrows():
                    alerts.append({
                        'نوع_التنبيه': 'مخزون منخفض',
                        'اسم_العنصر': item['اسم_العنصر'],
                        'الكمية_الحالية': item['الكمية_الحالية'],
                        'مستوى_الخطر': 'عالي' if item['الكمية_الحالية'] <= 5 else 'متوسط',
                        'التاريخ': datetime.now().strftime('%Y/%m/%d')
                    })
            
            return alerts
            
        except Exception as e:
            print(f"خطأ في الحصول على التنبيهات: {e}")
            return []
    
    def _get_comprehensive_statistics(self, project_name):
        """إحصائيات شاملة"""
        try:
            stats = {}
            
            # إحصائيات المخزون
            inventory_df = self.excel_manager.get_inventory_summary(project_name)
            if not inventory_df.empty:
                stats['إجمالي العناصر'] = len(inventory_df)
                stats['إجمالي الكميات'] = int(inventory_df['الكمية_الحالية'].sum())
                stats['متوسط الكمية لكل عنصر'] = round(inventory_df['الكمية_الحالية'].mean(), 2)
                stats['أعلى كمية'] = int(inventory_df['الكمية_الحالية'].max())
                stats['أقل كمية'] = int(inventory_df['الكمية_الحالية'].min())
            
            # إحصائيات الحركات
            project_file = os.path.join("projects", f"{project_name}_Transactions.xlsx")
            if os.path.exists(project_file):
                try:
                    transactions_df = pd.read_excel(project_file, engine='openpyxl')
                    if transactions_df is not None and not transactions_df.empty:
                        stats['إجمالي الحركات'] = len(transactions_df)
                        stats['حركات الإدخال'] = len(transactions_df[transactions_df['نوع_العملية'] == 'إدخال'])
                        stats['حركات الإخراج'] = len(transactions_df[transactions_df['نوع_العملية'] == 'إخراج'])
                        
                        # التحقق من وجود التواريخ وتحويلها
                        if 'التاريخ' in transactions_df.columns:
                            transactions_df['التاريخ'] = pd.to_datetime(transactions_df['التاريخ'])
                            stats['آخر حركة'] = transactions_df['التاريخ'].max().strftime('%Y/%m/%d')
                            stats['أول حركة'] = transactions_df['التاريخ'].min().strftime('%Y/%m/%d')
                except Exception as e:
                    print(f"خطأ في تحليل الحركات: {e}")
                    stats['إجمالي الحركات'] = 0
            
            # معلومات التقرير
            stats['تاريخ إنشاء التقرير'] = datetime.now().strftime('%Y/%m/%d %H:%M:%S')
            stats['نسخة النظام'] = '2.0'
            stats['نوع التقرير'] = 'شامل ومفصل'
            
            return stats
            
        except Exception as e:
            print(f"خطأ في الإحصائيات الشاملة: {e}")
            return {'خطأ': 'فشل في جمع الإحصائيات'}
    
    def export_filtered_report(self, project_name, start_date, end_date, filepath):
        """تصدير تقرير مفلتر حسب التاريخ مع كل محتوى التقرير الشامل (12 ورقة)"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            if not filepath:
                filename = f"تقرير_مخصص_{project_name}_{timestamp}.xlsx"
                filepath = os.path.join(self.reports_path, filename)
            
            # الحصول على الحركات في الفترة المحددة
            project_file = os.path.join("projects", f"{project_name}_Transactions.xlsx")
            if not os.path.exists(project_file):
                return None, "لا توجد حركات مسجلة للمشروع"
            
            # قراءة جميع الحركات
            all_transactions = pd.read_excel(project_file, engine='openpyxl')
            if all_transactions is None or all_transactions.empty:
                return None, "لا توجد حركات في الفترة المحددة"
            
            # تحويل التاريخ
            all_transactions['التاريخ'] = pd.to_datetime(all_transactions['التاريخ'])
            
            # فلترة حسب التاريخ
            filtered_df = all_transactions[
                (all_transactions['التاريخ'].dt.date >= start_date) & 
                (all_transactions['التاريخ'].dt.date <= end_date)
            ]
            
            if filtered_df.empty:
                return None, f"لا توجد حركات في الفترة من {start_date} إلى {end_date}"
            
            # الحصول على المخزون الحالي
            inventory_df = self.excel_manager.get_inventory_summary(project_name)
            
            # إنشاء ملف Excel مع جميع الأوراق (مثل التقرير الشامل)
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                
                # 1. معلومات التقرير
                report_info = {
                    'المعلومة': ['المشروع', 'نوع التقرير', 'من تاريخ', 'إلى تاريخ', 'تاريخ الإنشاء', 'عدد الحركات'],
                    'القيمة': [
                        project_name,
                        'تقرير مخصص بالتاريخ',
                        start_date.strftime('%Y/%m/%d'),
                        end_date.strftime('%Y/%m/%d'),
                        datetime.now().strftime('%Y/%m/%d %H:%M'),
                        len(filtered_df)
                    ]
                }
                info_df = pd.DataFrame(report_info)
                info_df.to_excel(writer, sheet_name='معلومات التقرير', index=False)
                
                # 2. الملخص التنفيذي (للفترة المحددة)
                try:
                    summary_data = {
                        'البيان': [
                            'إجمالي العناصر المسجلة',
                            'إجمالي الكميات الحالية', 
                            'عناصر بمخزون منخفض',
                            'إجمالي الحركات في الفترة',
                            'حركات الإدخال',
                            'حركات الإخراج',
                            'تاريخ إنشاء التقرير',
                            'حالة المشروع'
                        ],
                        'القيمة': [
                            len(inventory_df) if inventory_df is not None else 0,
                            int(inventory_df['الكمية_الحالية'].sum()) if inventory_df is not None else 0,
                            len(inventory_df[inventory_df['الكمية_الحالية'] <= 10]) if inventory_df is not None else 0,
                            len(filtered_df),
                            len(filtered_df[filtered_df['نوع_العملية'] == 'إدخال']),
                            len(filtered_df[filtered_df['نوع_العملية'] == 'إخراج']),
                            datetime.now().strftime('%Y/%m/%d %H:%M'),
                            'نشط'
                        ]
                    }
                    exec_df = pd.DataFrame(summary_data)
                    exec_df.to_excel(writer, sheet_name='الملخص التنفيذي', index=False)
                except Exception as e:
                    print(f"خطأ في الملخص التنفيذي: {e}")
                
                # 3. المخزون الحالي
                try:
                    if inventory_df is not None and not inventory_df.empty:
                        enhanced_inventory = self._enhance_inventory_details(inventory_df)
                        enhanced_inventory.to_excel(writer, sheet_name='المخزون الحالي', index=False)
                except Exception as e:
                    print(f"خطأ في المخزون: {e}")
                
                # 4. جميع الحركات المفلترة
                try:
                    enhanced_df = self._enhance_transactions_details(filtered_df)
                    enhanced_df.to_excel(writer, sheet_name='الحركات', index=False)
                except Exception as e:
                    print(f"خطأ في الحركات: {e}")
                
                # 5. حركات الإدخال
                try:
                    incoming = filtered_df[filtered_df['نوع_العملية'] == 'إدخال'].copy()
                    if not incoming.empty:
                        enhanced_incoming = self._enhance_transactions_details(incoming)
                        enhanced_incoming.to_excel(writer, sheet_name='حركات الدخول', index=False)
                except Exception as e:
                    print(f"خطأ في حركات الدخول: {e}")
                
                # 6. حركات الإخراج
                try:
                    outgoing = filtered_df[filtered_df['نوع_العملية'] == 'إخراج'].copy()
                    if not outgoing.empty:
                        enhanced_outgoing = self._enhance_transactions_details(outgoing)
                        enhanced_outgoing.to_excel(writer, sheet_name='حركات الخروج', index=False)
                except Exception as e:
                    print(f"خطأ في حركات الخروج: {e}")
                
                # 7. تحليل حسب التصنيف
                try:
                    category_analysis = self._analyze_by_category(filtered_df)
                    if category_analysis is not None and not category_analysis.empty:
                        category_analysis.to_excel(writer, sheet_name='تحليل التصنيفات', index=False)
                except Exception as e:
                    print(f"خطأ في التصنيفات: {e}")
                
                # 8. التحليل الزمني
                try:
                    time_analysis = self._analyze_by_time(filtered_df)
                    if time_analysis is not None and not time_analysis.empty:
                        time_analysis.to_excel(writer, sheet_name='التحليل الزمني', index=False)
                except Exception as e:
                    print(f"خطأ في التحليل الزمني: {e}")
                
                # 9. أعلى العناصر حركة
                try:
                    top_items = self._get_top_items_analysis(filtered_df)
                    if top_items is not None and not top_items.empty:
                        top_items.to_excel(writer, sheet_name='أعلى العناصر حركة', index=False)
                except Exception as e:
                    print(f"خطأ في أعلى العناصر: {e}")
                
                # 10. قاعدة العناصر
                try:
                    master_items = self.excel_manager.get_all_items()
                    if master_items is not None and not master_items.empty:
                        master_items.to_excel(writer, sheet_name='قاعدة العناصر', index=False)
                except Exception as e:
                    print(f"خطأ في قاعدة العناصر: {e}")
                
                # 11. التنبيهات
                try:
                    alerts_data = self._get_alerts_data(project_name)
                    if alerts_data and len(alerts_data) > 0:
                        alerts_df = pd.DataFrame(alerts_data)
                        alerts_df.to_excel(writer, sheet_name='تنبيهات وإنذارات', index=False)
                except Exception as e:
                    print(f"خطأ في التنبيهات: {e}")
                
                # 12. الإحصائيات الشاملة
                try:
                    comprehensive_stats = self._get_comprehensive_statistics(project_name)
                    if comprehensive_stats and len(comprehensive_stats) > 0:
                        stats_df = pd.DataFrame([comprehensive_stats]).T
                        stats_df.columns = ['القيمة']
                        stats_df.index.name = 'الإحصائية'
                        stats_df.to_excel(writer, sheet_name='إحصائيات شاملة')
                except Exception as e:
                    print(f"خطأ في الإحصائيات: {e}")
                
                # تنسيق جميع الأوراق
                workbook = writer.book
                for sheet_name in writer.sheets:
                    sheet = writer.sheets[sheet_name]
                    self._format_excel_sheet(sheet, workbook)
            
            return filepath, "تم إنشاء التقرير المخصص بنجاح! يحتوي على 12 ورقة عمل مع جميع التفاصيل مفلترة حسب التاريخ"
            
        except Exception as e:
            return None, f"خطأ في إنشاء التقرير المخصص: {str(e)}"
    
    def _enhance_transactions_details(self, transactions_df):
        """تحسين تفاصيل الحركات"""
        try:
            if transactions_df.empty:
                return transactions_df
            
            enhanced_df = transactions_df.copy()
            
            # تحويل التاريخ للتأكد من التنسيق الصحيح
            enhanced_df['التاريخ'] = pd.to_datetime(enhanced_df['التاريخ'])
            
            # إضافة الشهر ويوم الأسبوع
            enhanced_df['الشهر'] = enhanced_df['التاريخ'].dt.strftime('%Y-%m')
            enhanced_df['يوم_الأسبوع'] = enhanced_df['التاريخ'].dt.day_name()
            
            # ترجمة أيام الأسبوع للعربية
            day_translation = {
                'Monday': 'الإثنين',
                'Tuesday': 'الثلاثاء', 
                'Wednesday': 'الأربعاء',
                'Thursday': 'الخميس',
                'Friday': 'الجمعة',
                'Saturday': 'السبت',
                'Sunday': 'الأحد'
            }
            enhanced_df['يوم_الأسبوع'] = enhanced_df['يوم_الأسبوع'].map(day_translation)
            
            # إعادة ترتيب الأعمدة
            columns_order = ['التاريخ', 'الشهر', 'يوم_الأسبوع', 'اسم_العنصر', 'التصنيف', 
                           'نوع_العملية', 'الكمية', 'المسؤول', 'ملاحظات']
            
            # الاحتفاظ فقط بالأعمدة الموجودة
            existing_columns = [col for col in columns_order if col in enhanced_df.columns]
            remaining_columns = [col for col in enhanced_df.columns if col not in existing_columns]
            final_columns = existing_columns + remaining_columns
            
            enhanced_df = enhanced_df[final_columns]
            
            # ترتيب حسب التاريخ (الأحدث أولاً)
            enhanced_df = enhanced_df.sort_values('التاريخ', ascending=False)
            
            return enhanced_df
            
        except Exception as e:
            print(f"خطأ في تحسين بيانات الحركات: {e}")
            return transactions_df
    